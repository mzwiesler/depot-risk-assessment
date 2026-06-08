import zipfile
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


def read_xlsx_file_without_styles(file_path: str) -> pd.DataFrame:
    """
    Reads the holdings Excel file from the downloads folder.

    Args:
        file_path: Optional path to the Excel file. If None, uses default path.

    Returns:
        pd.DataFrame: DataFrame containing the holdings data
    """
    src = Path(file_path)

    if not src.exists():
        raise FileNotFoundError(f"File not found: {src}")

    dst = src.with_name("nostyles.xlsx")

    with zipfile.ZipFile(src, "r") as zin:
        with zipfile.ZipFile(dst, "w") as zout:
            for item in zin.infolist():
                # skip the stylesheet
                if item.filename != "xl/styles.xml":
                    zout.writestr(item, zin.read(item.filename))

    try:
        wb = load_workbook(
            dst,
            read_only=True,
            data_only=True,
        )

        ws = wb.active
        if ws is None:
            raise ValueError(f"No active worksheet found in {src}")

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
    finally:
        dst.unlink(missing_ok=True)

    logger.info("Successfully read %s rows from %s", len(data), src.name)
    return pd.DataFrame(data)


if __name__ == "__main__":
    path = (
        "downloads/Fondszusammensetzung_Amundi S&P World Health Care"
        " Screened UCITS ETF Acc_IE0006FM6MI8_07_01_2026.xlsx"
    )
    holdings_df = read_xlsx_file_without_styles(path)
    logger.info("DataFrame shape: %s", holdings_df.shape)
    logger.info("DataFrame info:")
    holdings_df.to_csv("downloads/holdings.csv", index=False, header=False)
    logger.info("Wrote CSV to downloads/holdings.csv")
